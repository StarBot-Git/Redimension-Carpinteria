from app.config import settings
from pathlib import Path
import pandas as pd
import shutil

class TopBarController():
    def __init__(self):
        self.project = None

    def cotizar(self):
        dir_BreakdownsProject_Folder = f"{settings.ONEDRIVE_PROJECTS_DIR}\\{self.project}\\Despieces"
        dir_quote_template = r"assets\templates\cotizador.xlsm"
        dir_quote_outFile = f"{settings.ONEDRIVE_PROJECTS_DIR}\\{self.project}"

        list_breakdowns = Path(dir_BreakdownsProject_Folder).iterdir()

        df_list = []
        
        for breakdown_Path in list_breakdowns:
            if "despiece" in str(breakdown_Path).lower():
                try:
                    # Leer el archivo Excel y agregarlo a la lista
                    df_temp = pd.read_excel(breakdown_Path)
                    
                    # Eliminar la columna "No." si existe
                    if "No." in df_temp.columns:
                        df_temp = df_temp.drop(columns=["No."])
                    
                    df_list.append(df_temp)
                    print(f"Archivo leído: {breakdown_Path.name}")
                except Exception as e:
                    print(f"Error al leer {breakdown_Path.name}: {e}")
        
        # Concatenar todos los dataframes
        if df_list:
            df_combined = pd.concat(df_list, ignore_index=True)
            
            # Copiar el template a la carpeta de salida
            template_name = Path(dir_quote_template).name
            output_file = Path(dir_quote_outFile) / template_name
            
            shutil.copy2(dir_quote_template, output_file)
            print(f"Template copiado a: {output_file}")
            
            # Usar openpyxl directamente para mantener las macros
            from openpyxl import load_workbook
            
            wb = load_workbook(output_file, keep_vba=True)
            
            # Verificar si existe la hoja "Despiece"
            if 'Despiece' in wb.sheetnames:
                ws = wb['Despiece']
                
                # Limpiar solo el contenido de las celdas desde B2 en adelante
                # Mantener la columna A y la fila 1 intactas
                for row in ws.iter_rows(min_row=2, min_col=2):
                    for cell in row:
                        cell.value = None
            else:
                # Si no existe, crearla
                ws = wb.create_sheet('Despiece')
            
            # Escribir encabezados desde B1
            for col_idx, col_name in enumerate(df_combined.columns, 2):  # Empieza en columna 2 (B)
                ws.cell(row=1, column=col_idx, value=col_name)
            
            # Escribir datos desde B2
            for row_idx, row_data in enumerate(df_combined.itertuples(index=False), 2):  # Empieza en fila 2
                for col_idx, value in enumerate(row_data, 2):  # Empieza en columna 2 (B)
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Guardar y cerrar
            wb.save(output_file)
            wb.close()
            
            print(f"Datos escritos en la hoja 'Despiece' de {output_file}")
            return output_file
        else:
            print("No se encontraron archivos de despiece para procesar")
            return None


    @staticmethod
    def seccionar():
        print("🧩 Seccionando...")

    @staticmethod
    def informacion():
        print("ℹ️ Mostrando información...")