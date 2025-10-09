import os
import win32com.client
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

from PySide6.QtWidgets import (
    QFrame, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QTableView, QSizePolicy, QHeaderView, QStyledItemDelegate
)
from PySide6.QtGui import QStandardItemModel, QStandardItem
from PySide6.QtCore import Qt, QSize
from PySide6.QtGui import QIcon

from app.config import settings
from app.ui.widgets.table_check import CheckDelegate


class MetricsEditorView(QFrame):
    # ======== Headers | Tabla metrica ========
    HEADERS_METRICS = ["Parametro", "Valor", "Unidad"]
    HEADERS_PROPS = ["Pieza", "Material", "Texturizado", "Material Canto", "A1", "A2", "L1", "L2"]

    # ======== Parametros de salida | Despiece ========
    PARAMETROS_SALIDA = ["Ancho", "Largo", "Espesor"]

    # ======== Proyecto actual ========
    NOMBRE_PROYECTO = None

    KEY_FIELD = "Pieza"

    """
        MetricsEditorView():
            - Titulo de modelo
            - Boton 'modo de tabla': Metricas o Propiedades
            - Barra de busqueda
            - Boton 'Ver Modelo'
            - Boton 'Planos'
            - Boton 'Despiece CSV'
            - Boton 'Cargar cambios'
            - Boton 'Guardar cambios'
            - Tabla de valores | Metricas o Propiedades del modelo
    """
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setObjectName("MetricsCard")

        # ======== Elemento raiz ========

        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(12)

        # ======== Contenedor superior ========

        root_sup = QHBoxLayout()
        root_sup.setContentsMargins(16, 16, 16, 16)

        # --- Titulo de modelo ----
        self.title = QLabel("Editor de metricas")
        self.title.setObjectName("MetricsCardTitle")
        
        # --- Boton | Modo metricas o propiedades ---
        self.btn_Table = QPushButton("Metricas")
        self.btn_Table.clicked.connect(self.toggle_table)
        self.btn_Table.setProperty("State", True) # True -> Metricas
        self.set_Button_Style(self.btn_Table, icon_path="assets/icons/toggle_State1.svg")
        
        root_sup.addWidget(self.title)
        root_sup.addStretch(1)
        root_sup.addWidget(self.btn_Table)

        root.addLayout(root_sup)

        # ======== Contenedor | buscador ========

        search_row = QHBoxLayout()
        search_row.setSpacing(8)

        # --- Barra buscadora de metricas ---

        self.search = QLineEdit()
        self.search.setPlaceholderText("Buscar metrica...")
        self.search.setObjectName("SearchField")
        search_row.addWidget(self.search)

        root.addLayout(search_row)

        # ======== Botones | Funcionalidades varias ========

        actions = QHBoxLayout()

        # --- Boton | Ver Modelo ---
        self.btn_Model   = QPushButton(" Ver Modelo")
        self.btn_Model.clicked.connect(self.on_view_model)
        self.set_Button_Style(self.btn_Model, icon_path="assets/icons/eye.svg")

        # --- Boton | Planos ---
        self.btn_Drawings = QPushButton(" Planos")
        self.btn_Drawings.clicked.connect(self.on_view_drawings)
        self.set_Button_Style(self.btn_Drawings, icon_path="assets/icons/blueprint.svg")

        # --- Boton | Despiece CSV ---
        self.btn_Import= QPushButton(" Despiece CSV")
        self.btn_Import.clicked.connect(self.on_export_csv)
        self.set_Button_Style(self.btn_Import, icon_path="assets/icons/download.svg")

        # --- Boton | Cargar cambios ---
        self.btn_Load = QPushButton(" Cargar cambios")
        self.btn_Load.clicked.connect(self.on_load_changes)
        self.set_Button_Style(self.btn_Load, icon_path="assets/icons/upload.svg")

        # --- Boton | Guardar Cambios ---
        self.btn_Save= QPushButton(" Guardar Cambios")
        self.btn_Save.clicked.connect(self.on_save_changes)
        self.set_Button_Style(self.btn_Save, icon_path="assets/icons/save.svg")

        actions.addWidget(self.btn_Model)
        actions.addWidget(self.btn_Drawings)
        actions.addWidget(self.btn_Import)
        actions.addStretch(1)
        actions.addWidget(self.btn_Load)
        actions.addWidget(self.btn_Save)

        root.addLayout(actions)

        # ======== Tabla de metricas ========

        self.rows_metrics = [] # Metricas del modelo actual
        self.rows_props = [] # Propiedades del modelo actual

        self.view = QTableView(self)
        self.view.setObjectName("MetricsTable")
        self.view.setAlternatingRowColors(True) # Alternar colores de filas
        self.view.setSelectionBehavior(QTableView.SelectRows) # Seleccionar filas completas
        self.view.setSortingEnabled(False) # Deshabilitar permisos de ordenamiento
        root.addWidget(self.view, 1)

        # - Modelo | Tabla metrica -

        self.model = QStandardItemModel(0, len(self.HEADERS_METRICS), self) # 0 filas, n columnas
        self.model.setHorizontalHeaderLabels(self.HEADERS_METRICS) # Encabezados de columnas
        self.view.setModel(self.model)

        # Oculta header vertical y quita boton de esquina
        self.view.verticalHeader().setVisible(False)
        self.view.setCornerButtonEnabled(False)

        # La vista debe poder expandirse
        self.view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.view.setMouseTracking(True)

        # Aplicar estilo base
        self._apply_header_bounds(True)
    
    
    """
        set_model_name():

        Actualiza el título del MetricsEditorView con el nombre del modelo seleccionado.
    """
    def set_model_name(self, model_name: str):
        self.model_name = model_name

        if model_name:
            self.title.setText(f"Editor de metricas - {model_name}")
        else:
            self.title.setText("Editor de metricas")

    # ---------- Buttons ----------

    """
        set_Button_Style():
            Estilo base de los botones relacionados a la tabla
    """
    def set_Button_Style(self, button: QPushButton, icon_path: str = ""):

        # === Estilo base | Button ===
        button.setObjectName("Button_Style")
        button.setEnabled(False)
        button.setMinimumHeight(32)
        button.setCursor(Qt.PointingHandCursor)

        if icon_path:
            button.setIcon(QIcon(icon_path)) 
            button.setIconSize(QSize(16, 16))  

    """
        on_view_model():
            Funcion encargada de abrir visualmente el modelo actual desde Inventor
    """
    def on_view_model(self):

        # === Verificacion | assembler document ===
        folder = Path(self.model_path)
        iam_files = list(folder.glob("*.iam"))

        if iam_files:
            asm_path = str(iam_files[0])  # primer .iam que hay
            print(f"✅ Ensamble encontrado: {asm_path}")
        else:
            asm_path = None
            print("❌ No se encontró ningún archivo .iam")
        
        if os.path.exists(asm_path):
            self.asmDoc = self.inventor.Documents.Open(asm_path)
            print("✅ Ensamble CUERPO abierto")
        else:
            print("❌ No se encontró el archivo de ensamble")
            return
        
        # === Mostrar Inventor ===
        self.inventor.Visible = True

    """
        on_view_drawings():
            Funcion que permite generar los planos de un modelo
    """
    def on_view_drawings(self):
        print("Ver planos")
        # FUNCION PENDIENTE

    """
        on_import_csv():
            Funcion encargada de exportar el despiece en formato Excel
    """
    def on_export_csv(self):
        data = [] # Tabla de salida

        # === Verificar | Ensamble abierto ===
        folder = Path(self.model_path)
        iam_files = list(folder.glob("*.iam"))

        if iam_files:
            asm_path = str(iam_files[0])  # primer .iam que encuentre
            print(f"✅ Ensamble encontrado: {asm_path}")
        else:
            asm_path = None
            print("❌ No se encontró ningún archivo .iam")

        if os.path.exists(asm_path):
            self.asmDoc = self.inventor.Documents.Open(asm_path)
            print("✅ Ensamble abierto para importar")
        else:

            print("❌ No se encontró el archivo de ensamble")
            return
        
        # === Extraccion de parametros por pieza ===
        for occ in self.asmDoc.ComponentDefinition.Occurrences:
            partDoc = occ.Definition.Document
            nombre = partDoc.DisplayName

            print(f"📦 Componente: {nombre}")

            # Si: El documento actual no es una pieza, entonces, pase al siguiente
            if "ipt" not in nombre:
                continue

            nombre = nombre.split('.')[0] # Quitar el formato del documento.

            fila = {"Designación": nombre} # Columna | Nombre
            fila["Tipo"] = "Tablero" # Columna | Tipo

            # ====== Busqueda | Pieza en la tabla de propiedades ======
            pieza_actual = None

            for row in self.rows_props:
                if row.get("Pieza") in nombre.split('.')[0]:

                    pieza_actual = row
                    print(f"✅ ES: {row.get("Pieza")}")

                    break

            # ====== Adjuntar | Informacion de la propiedades de la pieza ======
            if pieza_actual:
                fila["Material"] = pieza_actual.get("Material")
                fila["Material Canto"] = pieza_actual.get("Material Canto")
                fila["A1"] = pieza_actual.get("A1")
                fila["A2"] = pieza_actual.get("A2")
                fila["L1"] = pieza_actual.get("L1")
                fila["L2"] = pieza_actual.get("L2")
                fila["Texturizado"] = pieza_actual.get("Texturizado")

            # ====== Adjuntar | Parametros de la pieza ======
            try:
                params = partDoc.ComponentDefinition.Parameters

                for clave in self.PARAMETROS_SALIDA:
                    try:
                        fila[clave] = params.Item(clave).Value*10  # Convertir cm a mm
                    except:
                        fila[clave] = None  # Si no existe, se deja vacio
            except:
                # Si no hay parámetros, se deja todo en None
                for clave in self.PARAMETROS_SALIDA:
                    fila[clave] = None
            
            data.append(fila)

        # === Conversion | Dataframe ===

        df = pd.DataFrame(data)
        df = self.breakdown_CSV_Format(df=df)

        # === Exportar | Despiece ===

        output_csv = f"{settings.ONEDRIVE_PROJECTS_DIR}\\{self.NOMBRE_PROYECTO}\\Despieces\\despiece-{self.model_name}.xlsx"

        wb = load_workbook(r"assets\templates\despiece.xlsx")
        ws = wb["Despiece"] # Hoja donde esta el diseño

        start_row, start_col = 1, 2   # B1

        # Escribir encabezados
        for j, col in enumerate(df.columns, start=start_col):
            ws.cell(row=start_row, column=j).value = col
        # Escribir filas
        for i, row in enumerate(df.itertuples(index=False), start=start_row+1):
            for j, val in enumerate(row, start=start_col):
                ws.cell(row=i, column=j).value = val

        wb.save(output_csv)
        print(f"Archivo generado en: {output_csv}")

    """
        on_save_changes():
            Guardar los cambios realizados al modelo actual.
    """
    def on_save_changes(self):
        if hasattr(self, "asmDoc") and self.asmDoc is not None:
            self.inventor.SilentOperation = True  
            self.asmDoc.Save()

            for ref in self.asmDoc.ReferencedDocuments:
                try:
                    if ref.Dirty:
                        ref.Save()
                except Exception as e:
                    print(f"No se pudo guardar {ref.FullFileName}: {e}")
            print("💾 Cambios guardados en ensamble y dependencias")

            self.inventor.SilentOperation = False

    """
        on_load_changes():
            Actualizar el modelo con los cambios aplicados.
    """
    def on_load_changes(self):
        # Asegurarnos que el ensamble está abierto
        folder = Path(self.model_path)
        iam_files = list(folder.glob("*.iam"))

        if iam_files:
            asm_path = str(iam_files[0])  # primer .iam que encuentre
            print(f"✅ Ensamble encontrado: {asm_path}")
        else:
            asm_path = None
            print("❌ No se encontró ningún archivo .iam")

        if not hasattr(self, "asmDoc") or self.asmDoc is None:
            if os.path.exists(asm_path):
                self.asmDoc = self.inventor.Documents.Open(asm_path)
                print("✅ Ensamble abierto para guardar")
            else:
                print("❌ No se encontró el archivo de ensamble")
                return

        if self.skeleton_doc:
            params = self.skeleton_doc.ComponentDefinition.Parameters
            parametros = self.get_param_dict()

            for nombre, valor in parametros.items():
                try:
                    params.Item(nombre).Value = valor/10  # Convertir mm a cm
                    print(f"✅ Parámetro '{nombre}' actualizado a {valor}")
                except Exception as e:
                    print(f"❌ Error en '{nombre}': {e}")

            # Propagar cambios
            self.skeleton_doc.Update()
            self.asmDoc.Update()

            # === Recalcular rows_metrics (solo memoria, no toca la tabla) ===
            try:
                all_params = [p for p in params]  # forzar iterable COM
                rows = []
                for p in all_params:
                    nombre_p = getattr(p, "Name", "") or ""
                    valor_p  = getattr(p, "Value", None)
                    unidad_p = getattr(p, "Units", "") or ""
                    expr_p   = getattr(p, "Expression", "") or ""

                    # Ignorar parámetros dependientes
                    if any(other.Name in expr_p and other.Name != nombre_p for other in all_params):
                        continue

                    # Convertir a mm (tú lees *10)
                    if isinstance(valor_p, (int, float)):
                        valor_p = valor_p * 10

                    rows.append({
                        "Parametro": str(nombre_p),
                        "Valor": valor_p,
                        "Unidad": unidad_p,
                    })

                self.rows_metrics = rows
                # print(f"🔁 rows_metrics actualizado: {len(rows)} filas")
            except Exception as e:
                print(f"⚠ No se pudo reconstruir rows_metrics: {e}")

    """
        toggle_table():
            Cambiar el tipo de tabla que se esta visualizando.
    """
    def toggle_table(self):
        # -- Extraer | Estado actual de la tabla --
        current_state = bool(self.btn_Table.property("State")) # True -> Metricas | False -> Propiedades

        new_state = not current_state # Toggle del estado actual

        # -- Cambio | Estado y contenido de la tabla --
        self.btn_Table.setProperty("State", new_state)
        print(f"Nuevo estado: {new_state}")

        self.set_TableData(new_state)

        icon_path = (
        "assets/icons/toggle_State1.svg" if new_state
        else "assets/icons/toggle_State2.svg" 
        )
        button_text = (
            "Metricas" if new_state else "propiedades"
        )
        self.btn_Table.setIcon(QIcon(icon_path))
        self.btn_Table.setText(button_text)

        # === Forzar refresco de estilo ===
        self.btn_Table.style().unpolish(self.btn_Table)
        self.btn_Table.style().polish(self.btn_Table)
        self.btn_Table.update()

    # ---------- Breakdown Format ----------

    """
        breakdown_CSV_Format():
            Funcion encargada de formatear el Dataframde para el despiece
    """
    def breakdown_CSV_Format(self, df: pd.DataFrame):
        # == Agrupar ==
        df_agrupado = df.groupby('Designación', as_index=False).agg({
            'Ancho': 'first',
            'Largo': 'first', 
            'Espesor': 'first',
            'Material': 'first',
            'Material Canto': 'first',
            'A1': 'first',
            'A2': 'first',
            'L1': 'first',
            'L2': 'first',
            'Texturizado': 'first'
        })

        # == Convertir a numero o NaN ==
        for col in ['Ancho', 'Largo', 'Espesor', 'A1', 'A2', 'L1', 'L2']:
            df_agrupado[col] = pd.to_numeric(df_agrupado[col], errors='coerce')

        # == Espesor en metros ==
        df_agrupado['Espesor'] = df_agrupado['Espesor'] / 1000.0

        # == Contar cantidad de piezas ==
        df_agrupado['Cantidad'] = df.groupby('Designación').size().values

        # == Area Final en m² ==
        mask_area = df_agrupado['Largo'].gt(0) & df_agrupado['Ancho'].gt(0)

        df_agrupado['Area - final'] = ''
        df_agrupado.loc[mask_area, 'Area - final'] = (((df_agrupado.loc[mask_area, 'Largo']/1000) * (df_agrupado.loc[mask_area, 'Ancho']/1000)).round(2).astype(str) + ' m²')

        # == Agregar columna Tipo = "Tablero" ==
        df_agrupado['Tipo'] = 'Tablero'

        # == Tabla principal organizada ==
        df_principal = df_agrupado[['Designación', 'Cantidad', 'Largo', 'Ancho', 'Espesor', 
                                    'Area - final', 'Tipo', 'Material', 'A1', 'A2', 'L1', 'L2', 'Texturizado']].copy()

        # == Verificador NaN ==        
        def safe_num(x, default=0.0):
            return float(x) if pd.notna(x) else default

        # == Filas de cantos ==
        filas_cantos = []
        for _, row in df_agrupado.iterrows():
            nombre_pieza = row['Designación']

            cantidad_pieza = int(safe_num(row['Cantidad'], 0))

            ancho = safe_num(row['Ancho'], 0.0)
            largo = safe_num(row['Largo'], 0.0)
            espesor = safe_num(row['Espesor'], 0.0)

            material_canto = row.get('Material Canto')
            if not isinstance(material_canto, str) or not material_canto.strip():
                material_canto = 'No especificado'

            # A1/A2/L1/L2 pueden ser NaN -> False
            tiene_A1 = safe_num(row['A1'], 0) == 1.0
            tiene_A2 = safe_num(row['A2'], 0) == 1.0
            tiene_L1 = safe_num(row['L1'], 0) == 1.0
            tiene_L2 = safe_num(row['L2'], 0) == 1.0

            # Ancho del canto:
            ancho_canto = 0

            if espesor >= 0 and espesor <= 0.015:
                ancho_canto = 0.019
            elif espesor > 0.015 and espesor <= 0.019:
                ancho_canto = 0.022
            elif espesor > 0.019 and espesor <= 0.030:
                ancho_canto = 0.033
            elif espesor > 0.030 and espesor <= 0.036:
                ancho_canto = 0.044
            elif espesor > 0.036 or espesor == 0.050:
                ancho_canto = 0.055

            # Cantos A
            if tiene_A1 or tiene_A2:
                cantidad_cantos_A = (int(tiene_A1) + int(tiene_A2)) * cantidad_pieza
                largo_canto_A = ancho
                filas_cantos.append({
                    'Designación': f"{nombre_pieza} - A",
                    'Cantidad': cantidad_cantos_A,
                    'Largo': largo_canto_A,
                    'Ancho': ancho_canto,
                    'Espesor': 0,
                    'Area - final': '',
                    'Tipo': 'Canto',  # ← Aquí está!
                    'Material': material_canto,
                    'A1': None,
                    'A2': None,
                    'L1': None,
                    'L2': None,
                    'Texturizado': None
                })

            # Cantos L
            if tiene_L1 or tiene_L2:
                cantidad_cantos_L = (int(tiene_L1) + int(tiene_L2)) * cantidad_pieza
                largo_canto_L = largo
                filas_cantos.append({
                    'Designación': f"{nombre_pieza} - L",
                    'Cantidad': cantidad_cantos_L,
                    'Largo': largo_canto_L,
                    'Ancho': ancho_canto,
                    'Espesor': 0,
                    'Area - final': '',
                    'Tipo': 'Canto',  # ← Aquí también!
                    'Material': material_canto,
                    'A1': None,
                    'A2': None,
                    'L1': None,
                    'L2': None,
                    'Texturizado': None
                })

        df_cantos = pd.DataFrame(filas_cantos)

        # === 5) Combinar ===
        df_final = pd.concat([df_principal, df_cantos], ignore_index=True)
        return df_final

    # ---------- Table Data Handling ----------

    """
        _parse_number():
            Funcion para convertir a dato flotante
    """
    def _parse_number(self, v):
        """Convierte el valor de la celda a float si se puede (soporta '1,23')."""
        if isinstance(v, (int, float)):
            return float(v)
        
        s = str(v).strip()
        if not s:
            return None
        
        try:
            return float(s.replace(",", "."))  # 1,23 -> 1.23
        except ValueError:
            return None

    """
        get_param_dict():
            Devuelve { PARAMETRO: VALOR } para usarlo igual que los 'parametros'.
            Si un valor no es numérico o está vacío, deja None.
    """
    def get_param_dict(self) -> dict[str, float | None]:
        
        d = {}
        for r in range(self.model.rowCount()):
            name_item = self.model.item(r, 0)
            val_item  = self.model.item(r, 1)
            if not name_item:
                continue
            name = name_item.text().strip()
            val  = self._parse_number(val_item.text() if val_item else "")
            if name:
                d[name] = val
        return d
    
    # ---------- Table Setup ----------

    """
        _apply_header_bounds():
            Funcion que estiliza el Header de la tabla
    """
    def _apply_header_bounds(self, is_metrics: bool):
        h = self.view.horizontalHeader()
        h.setStretchLastSection(False)
        h.setSectionsMovable(True)
        h.setMinimumSectionSize(30)
        h.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)

        if is_metrics:
            # Métricas: ["Parametro","Valor","Unidad"]

            # ------ Parametro | Se estira ------
            h.setSectionResizeMode(0, QHeaderView.Stretch)
            self.view.setItemDelegateForColumn(0, QStyledItemDelegate(self.view))

            h.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # Valor al contenido
            # opcional: dar un poquito más a "Valor"
            w = max(80, h.sectionSize(1) + 100)
            h.resizeSection(1, w)
            h.setSectionResizeMode(2, QHeaderView.Fixed)             # Unidad fijo
            self.view.setColumnWidth(2, 72)
        else:
            # Propiedades: ["Editar","Pieza","Material","Texturizado","Material Canto","A1","A2","L1","L2"]

            # ------ checkbox | Angosto ------
            h.setSectionResizeMode(0, QHeaderView.Fixed)             
            self.view.setItemDelegateForColumn(0, CheckDelegate(self.view))
            self.view.setColumnWidth(0, 28)

            # ------ Pieza | Se estira ------
            h.setSectionResizeMode(1, QHeaderView.Stretch)

            # ------ Material | Se adapta al contenido ------
            h.setSectionResizeMode(2, QHeaderView.ResizeToContents)

            # ------ Texturizado | Se adapta al contenido ------
            h.setSectionResizeMode(3, QHeaderView.ResizeToContents)

            # ------ Material Canto | Se adapta al contenido ------
            h.setSectionResizeMode(4, QHeaderView.ResizeToContents)

            # ------ A1,A2,L1,L2 | Angostos ------
            for col in (5, 6, 7, 8):
                h.setSectionResizeMode(col, QHeaderView.Fixed)
                self.view.setColumnWidth(col, 36)


    """
        set_TableData():
            Funcion que actualiza la tabla segun su modo(Metricas o Propiedades).
    """
    def set_TableData(self, state: bool):
        # === Seleccion | Metricas (True) o Propiedades (False) ===
        if state:
            headers = self.HEADERS_METRICS                      # ["Parametro","Valor","Unidad"]
            rows = self.rows_metrics
        else:
            # Para Propiedades, recuerda que ya agregaste la col "Editar" en el modelo
            # si usas la columna check. Si no, deja HEADERS_PROPS tal cual.
            headers = ["Editar"] + self.HEADERS_PROPS           # ["Editar","Pieza","Material",...]
            rows = self.rows_props

        self.model.removeRows(0, self.model.rowCount())
        self.model.setColumnCount(len(headers))
        self.model.setHorizontalHeaderLabels(headers)

        for r in rows:
            items = []
            if not state:
                # col 0 = checkbox (solo en Propiedades)
                chk = QStandardItem()
                chk.setCheckable(True)
                chk.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable)
                items.append(chk)

            # resto de columnas en orden de headers (saltando "Editar" si aplica)
            for hname in (headers[1:] if not state else headers):
                it = QStandardItem(str(r.get(hname, "")))

                it.setEditable(False)

                if hname == "Valor":
                    it.setEditable(True)

                items.append(it)

            self.model.appendRow(items)

        self._apply_header_bounds(state)   # <<-- una sola función

    # ---------- Inventor Model Integration ----------

    """
        load_inventor_model():
            Funcion encargada de cargar los datos(metricas y propiedades) del modelo seleccionado.
    """
    def load_inventor_model(self, model_path: str, loading_Bar):
        print(f"Cargando modelo de Inventor desde: {model_path}")

        # === Conexion con Inventor ===

        loading_Bar.set_Text("Cargando modelo de Inventor...")
        loading_Bar.set_Progress(4)

        self.inventor = win32com.client.Dispatch("Inventor.Application")
        self.inventor.Visible = False  # No mostrar la ventana de Inventor inicialmente

        loading_Bar.set_Text("Abriendo Skeleton...")
        loading_Bar.set_Progress(5)

        # === Apertura Skeleton Part ===
        self.model_path = model_path
        skeleton_path = f"{model_path}\\Skeleton Part.ipt"

        if os.path.exists(skeleton_path):
            print(f"✅ Skeleteon abierto")
            self.skeleton_doc = self.inventor.Documents.Open(skeleton_path)
        else:
            print(f"❌ No se encontró el archivo Skeleton del modelo")
            return
        
        loading_Bar.set_Text("Extrayendo parametro...")
        loading_Bar.set_Progress(6)

        params = self.skeleton_doc.ComponentDefinition.Parameters       

        print("Parámetros del modelo:")
        for param in params:
            print(f"- {param.Name}: {param.Value} {param.Units}")

        # ==== Fila de metricas ====
        rows = []
        for p in params:
            nombre = getattr(p, "Name", None) or getattr(p, "name", "")
            valor  = getattr(p, "Value", None) or getattr(p, "value", "")
            unidad = getattr(p, "Units", None) or getattr(p, "units", "")

            if any(other.Name in p.Expression  and other.Name != param.Name for other in params):
                print(f"Parametro '{nombre}' depende de otro, se ignora.")
                continue

            if isinstance(valor, (int, float)):
                valor = valor * 10

            rows.append({
                "Parametro": str(nombre),
                "Valor": valor,
                "Unidad": unidad,
            })

        self.rows_metrics = rows

        loading_Bar.set_Text("Extrayendo Propiedades de piezas...")
        loading_Bar.set_Progress(7)

        # ==== Fila de propiedades ====
            
        folder = Path(self.model_path)
        iam_files = list(folder.glob("*.iam"))

        if iam_files:
            asm_path = str(iam_files[0])  # primer .iam que encuentre
            print(f"✅ Ensamble encontrado: {asm_path}")
            self.rows_props = self.extract_properties_table_from_assembly(asm_path)
        else:
            asm_path = None
            print("❌ No se encontró ningún archivo .iam")
            self.rows_props = []
    
    # ---------- Inventor | Propiedades del ensamble ----------

    """
        faces_by_tag():
    """
    def faces_by_tag(sel, doc, tag: str):
        """Devuelve lista de entidades (faces/proxies) con iLogicEntityName == tag en un PartDocument."""
        am = doc.AttributeManager
        try:
            objs = am.FindObjects("iLogicEntityNameSet", "iLogicEntityName", tag)
            return [o for o in objs]  # puede ser vacío
        except:
            # Fallback: buscar por set+attr y filtrar por value
            try:
                objs = am.FindObjects("iLogicEntityNameSet", "iLogicEntityName")
            except:
                return []
            res = []
            for o in objs:
                try:
                    val = o.AttributeSets.Item("iLogicEntityNameSet").Item("iLogicEntityName").Value
                    if str(val) == tag:
                        res.append(o)
                except:
                    pass
            return res

    """
        collect_part_paths():
    """
    def collect_part_paths(self, asm_doc):
        """Recorre el ensamble (recursivo) y devuelve rutas únicas de todas las piezas .ipt."""
        seen = set()
        def walk(doc):
            try:
                refs = doc.ReferencedDocuments
            except:
                refs = []
            for r in refs:
                try:
                    f = r.FullFileName
                except:
                    continue
                if not f:
                    continue
                ext = Path(f).suffix.lower()
                if ext == ".ipt":
                    seen.add(f)
                elif ext == ".iam":
                    walk(r)
        walk(asm_doc)
        return sorted(seen)

    """
        extract_props_from_part():
    """
    def extract_props_from_part(self, doc):
        """Extrae Material, Texturizado, Material Canto y banderas A1/A2/L1/L2 de un PartDocument ya abierto."""
        # Material
        try:
            material = doc.ComponentDefinition.Material.Name
        except:
            material = ""

        # User Defined Property: "Texturizado"
        texturizado = ""
        try:
            udp = doc.PropertySets.Item("Inventor User Defined Properties")
            texturizado = udp.Item("Texturizado").Value
        except:
            pass

        # User Defined Property: "Material Canto"
        mat_canto = ""
        try:
            udp = doc.PropertySets.Item("Inventor User Defined Properties")
            mat_canto = udp.Item("Material Canto").Value
        except:
            pass

        # Tags presentes
        tags = ["A1", "A2", "L1", "L2"]
        presentes = {t: (len(self.faces_by_tag(doc, t)) > 0) for t in tags}

        pieza = Path(doc.FullFileName).stem if getattr(doc, "FullFileName", "") else doc.DisplayName

        return {
            "Pieza": pieza,
            "Material": material,
            "Texturizado": texturizado or "NT",
            "Material Canto": mat_canto or "No especificado",
            "A1": 1 if presentes["A1"] else 0,
            "A2": 1 if presentes["A2"] else 0,
            "L1": 1 if presentes["L1"] else 0,
            "L2": 1 if presentes["L2"] else 0,
        }

    """
        extract_properties_table_from_assembly():
    """
    def extract_properties_table_from_assembly(self, asm_path: str) -> list[dict]:
        """
        Abre el .iam (si no está abierto), recorre todas las .ipt referenciadas y
        devuelve list[dict] con HEADERS_PROPS.
        """
        
        inv = self.inventor
        # No forzamos visible; déjalo como tengas tu instancia principal
        # inv.Visible = False

        docs = inv.Documents

        # Abrir assembly si no está abierto
        asm_doc = None
        opened_asm_here = False
        for i in range(1, docs.Count + 1):
            d = docs.Item(i)
            if getattr(d, "FullFileName", "").lower() == asm_path.lower():
                asm_doc = d
                break
        if asm_doc is None:
            asm_doc = docs.Open(asm_path)
            opened_asm_here = True

        # Recolectar rutas de piezas
        part_paths = self.collect_part_paths(asm_doc)

        rows = []
        opened_here = []  # docs que abrimos aquí para luego cerrarlos

        try:
            for p in part_paths:
                # Buscar si la pieza ya está abierta

                if "skeleton" in Path(p).name.lower():
                    print(f"↪ Omitida por filtro: {p}")
                    continue

                part_doc = None
                for i in range(1, docs.Count + 1):
                    d = docs.Item(i)
                    if getattr(d, "FullFileName", "").lower() == p.lower():
                        part_doc = d
                        break
                if part_doc is None:
                    part_doc = docs.Open(p)
                    opened_here.append(part_doc)

                rows.append(self.extract_props_from_part(part_doc))
        finally:
            # Cerrar piezas que abrimos aquí
            for d in opened_here:
                try:
                    d.Close(True)  # True = save changes; usa False si no quieres guardar
                except:
                    pass
            # Cerrar ensamblado si lo abrimos aquí
            if opened_asm_here:
                try:
                    asm_doc.Close(True)
                except:
                    pass

        return rows
